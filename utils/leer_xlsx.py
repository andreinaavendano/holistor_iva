import xlrd
import openpyxl

from utils.olvido_clave_data import OlvidoClaveData


import pandas as pd


class leer_xlsx:
    @staticmethod
    def get_xlsx_olvidoClave_data():
        #df = pd.read_excel("utils/olvido_clave_input_data.xlsx")
        #wb = xlrd.open_workbook(
        #    r'F:/Projects/Selenium/SeleniumPython/pythonProjectHolistorIVAWeb/utils/olvido_clave_input_data.xlsx')
        wb = openpyxl.load_workbook(r"F:/Projects/Selenium/SeleniumPython/pythonProjectHolistorIVAWeb/utils"
                                    r"/olvido_clave_input_data.xlsx")

        sheet = wb.worksheets[0]
        data = []

        for i in range(2, sheet.max_row+1):
            olvidoClave_data = OlvidoClaveData(sheet.cell(i, 1).value,  # tenant
                                               sheet.cell(i, 2).value)  # email

            data.append(olvidoClave_data)
        return data
